import json
import openpyxl
from openpyxl import Workbook

#open and read JSON file
with open('OPC.json') as jsonFile:
    json_data = json.load(jsonFile)
'''
# Extracting the value of "common.ALLTYPES_NAME" in channels
channelnames = [channel["common.ALLTYPES_NAME"] for channel in json_data["project"]["channels"]]

print('channelnames',channelnames)
print(len(channelnames))
'''

def write_to_excel(data, filename):
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write data to the worksheet
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # Save the workbook
    wb.save(filename)






def extract_ip_addresses(devices):
    ip_addresses = []
    for device in devices:
        if "servermain.DEVICE_ETHERNET_COMMUNICATIONS_IP" in device:
            ip_addresses.append(device["servermain.DEVICE_ETHERNET_COMMUNICATIONS_IP"])
    return ip_addresses

def write_to_excel(json_data):
    # Create a new Workbook
    wb = Workbook()


    # Iterate over the range
    for i in range(453):
        # Create a new sheet for each iteration
        sheet = wb.create_sheet(title=f"Channel_{i + 1}")

        # Extract channel_name, device names, driver names, and IP addresses

        channel_name = json_data["project"]["channels"][i]["common.ALLTYPES_NAME"]
        devices = json_data["project"]["channels"][i]["devices"]
        device_names = [device["common.ALLTYPES_NAME"] for device in devices]
        driver_names = [device["servermain.MULTIPLE_TYPES_DEVICE_DRIVER"] for device in devices]
        ip_addresses = extract_ip_addresses(devices)

        # Write column names
        sheet['A1'] = 'channel Name'
        sheet['B1'] = 'Device Name'
        sheet['C1'] = 'Driver Name'
        sheet['D1'] = 'IP Address'

        # Write device names, driver names, and IP addresses into the sheet
        for row, (device_name, driver_name, ip_address) in enumerate(zip(device_names, driver_names, ip_addresses), start=2):
            sheet[f'A{row}'] = channel_name
            sheet[f'B{row}'] = device_name
            sheet[f'C{row}'] = driver_name
            sheet[f'D{row}'] = ip_address

    # Save the workbook
    wb.save("final.xlsx")

# Call the function on JSON data
write_to_excel(json_data)

'''
for i in range(453):
    device_names = [device["common.ALLTYPES_NAME"] for device in json_data["project"]["channels"][i]["devices"]]
    driver_names = [device["servermain.MULTIPLE_TYPES_DEVICE_DRIVER"] for device in json_data["project"]["channels"][i]["devices"]]
    print('devices_names of channle',i,device_names)
    print('driver_names', i, driver_names)
    for device in json_data["project"]["channels"][i]["devices"]:
        if "servermain.DEVICE_ETHERNET_COMMUNICATIONS_IP" in device:
            IP_address = device["servermain.DEVICE_ETHERNET_COMMUNICATIONS_IP"]
            print('IP_address', i, IP_address)
        else:
            # Handle the case when the key is missing
            print("Key 'servermain.DEVICE_ETHERNET_COMMUNICATIONS_IP' not found for device.")


    write_to_excel([device_names], filename)


print(f"Data written to {filename} successfully.")


# Example usage:



import re

def extract_info(channel_name):
    # Use regular expression to extract the information
    match = re.search(r'.*_.*_(\w|\d+)', channel_name)
    if match:
        return match.group(1)
    else:
        return ' '  # Return whitespace if no match found
for i in range(len(channelnames)):
     print(extract_info(channelnames[i]))

import re

def extract_numbers_from_list(input_list):
    numbers = []
    pattern = r'\d+'
    for item in input_list:
        match = re.findall(pattern, item)
        if match:
            numbers.append(''.join(match))
    return numbers

# Example usage:

print(extract_numbers_from_list(device_names))

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
sheet = workbook.active
# Write the channel names to the Excel sheet
for index, channel in enumerate(channelnames, start=1):
    sheet.cell(row=index, column=1, value=channel)
# Save the workbook
workbook.save("channel_names.xlsx")

def divide_by_prefix(channel_names):
    divided_channels = defaultdict(list)

    for channel in channel_names:
        prefix = channel.split('_')[0]
        divided_channels[prefix].append(channel)

    return divided_channels.items()

print(divide_by_prefix(channelnames))
'''
import re

'''
def divide_by_subtext(channel_names):
    divided_channels = defaultdict(list)

    for channel in channel_names:
        # Extract the subtext after the second underscore
        match = re.search(r'^[^_]+_[^_]+_(.*)', channel)
        if match:
            subtext = match.group(1)
            divided_channels[subtext].append(channel)

    return divided_channels
print(divide_by_subtext(channelnames))
'''







